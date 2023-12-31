from openpyxl.styles import Alignment
import openpyxl
from openpyxl.utils import get_column_letter
import os


def total_amount(worksheet):
    ws = worksheet
    # 获取最大行
    max_row = ws.max_row
    # 指定单元格范围
    ceel_range = ws['A2':f'Q{max_row + 1}']
    # 设置单元格居中和自动换行
    for i in ceel_range:
        for j in i:
            j.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    # 调整列宽
    #          A    B   C   D   E   F   G  H  I   J   K   L   M   N   O   P
    listnum = [12,  17, 11, 9,  11, 20, 8, 8, 8, 11, 13, 20,  26, 7, 20]
    for j in range(1, len(listnum) + 1):
        # 根据列的数字返回字母
        k = openpyxl.utils.get_column_letter(j)
        # 根据listnum列表设置列宽
        ws.column_dimensions[k].width = listnum[j - 1]
    # 调整行高
    for a in range(1, max_row+1):
        ws.row_dimensions[a].height = 30
    # 指定字体样式
    font_style = openpyxl.styles.Font(name=u'宋体', bold=False, italic=False, size=11)
    # 指定边框样式
    side = openpyxl.styles.Side(style='thin', color='000000')
    border = openpyxl.styles.Border(left=side, right=side, top=side, bottom=side)
    # 循环读取单元格
    for row in worksheet[f'A3:O{max_row}']:
        for cell in row:
            # 设置单元格字体 调整字体
            cell.font = font_style
            # 设置单元格边框
            cell.border = border



# 获取当前文件夹路径
Current_Folder_path = os.getcwd()
# 遍历当前文件夹下的文件
file_list = os.listdir(Current_Folder_path)
# 提取后缀为xlsx的文件
xlsx_list = [a for a in file_list if a.endswith('.xlsx')]
# 循环提取xlsx文件
for xlsx in xlsx_list:
    file_name = xlsx.split('.')[0]+'-格式调整.xlsx'
    print(file_name)
    #  生成文件路径
    xlsx_path = os.path.join(Current_Folder_path, xlsx)
    # 使用openpyxl读取xlsx文件
    workbook = openpyxl.load_workbook(xlsx_path)
    # 循环读取工作表
    for work_sheet in workbook.sheetnames:
        # 循环调用调整表格格式函数
        total_amount(workbook[work_sheet])
        print(work_sheet)
    # 工作表另存
    workbook.save(os.path.join(Current_Folder_path, file_name))
